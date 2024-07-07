from openpyxl import load_workbook

file_path = r'C:\V2G\Schedule\data\grid_load.xlsx'
workbook =load_workbook(file_path)

sheet = workbook['Area1_Load']
column_count =sheet.max_column
print(f"列数为：{column_count}")

from openpyxl import load_workbook
from openpyxl import Workbook

# 打开原始 Excel 文件
input_file_path = r'C:\V2G\Schedule\data\grid_load.xlsx'
input_workbook = load_workbook(input_file_path)
input_sheet = input_workbook.active

# 创建新的 Excel 文件
output_workbook = Workbook()
output_sheet = output_workbook.active

# 将原始 Excel 文件的数据整理成一列
output_row = 1
for row in input_sheet.iter_rows():
    for cell in row:
        output_sheet.cell(row=output_row, column=1, value=cell.value)
        output_row += 1

# 保存新的 Excel 文件
output_file_path = r'C:\V2G\Schedule\data\grid_load_preprogressed.xlsx'
output_workbook.save(output_file_path)

print("Excel 文件已整理完成！")
